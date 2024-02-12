from openpyxl import load_workbook
import json

file = './xlsx_data/sem2.xlsx'
wb = load_workbook(file)
data = {}

# region SEMESTER 2
for x in wb.sheetnames:
    data[x] = {}
    start = False
    prev = ""
    prevsubject = ""
    for row in wb[x].iter_rows(min_row=2, max_row=wb[x].max_row-1, min_col=1, max_col=wb[x].max_column):
        if row[0].value is None and not start:
            continue

        if row[0].value is None and start:
            pass
        else:
            if row[0].value.startswith("Course"):
                start = True
                continue
        if start:

            prev = row[0].value or prev
            if prevsubject == "":
                prevsubject = row[1].value
            prevsubject = row[1].value or prevsubject

            if prev in data[x]:
                data[x][prev].append([prevsubject] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[2:] if cell.value is not None])
            else:
                data[x][prev] = [[prevsubject] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[2:] if cell.value is not None]]
# endregion


with open('./json_data/sem2.json', 'w') as f:
    json.dump(data, f)