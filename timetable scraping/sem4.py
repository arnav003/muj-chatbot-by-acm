from openpyxl import load_workbook
import json

file = './xlsx_data/sem4.xlsx'
wb = load_workbook(file)
data = {}

# region SEMESTER 4
for x in wb.sheetnames:
    data[x] = {}
    start = False
    prev = ""
    prevsubject = ""
    sectionname = str(x)
    for row in wb[x].iter_rows(min_row=2, max_row=wb[x].max_row-1, min_col=1, max_col=wb[x].max_column):
        if row[0].value is None and not start:
            continue

        if row[0].value is None and start:
            pass
        else:
            if row[0].value.startswith("Course"):
                start = True
                continue
        if start:
            prev = row[0].value or prev
            prevsubject = row[1].value or prevsubject
            
            if row[5].value == None:
                if prev in data[x]:
                    data[x][prev].append([prevsubject] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[2:] if cell.value is not None])
                else:
                    data[x][prev] = [[prevsubject] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[2:] if cell.value is not None]]
            else:
                if row[5].value == sectionname + "2":
                    continue
                
                if prev in data[x]:
                    data[x][prev][sectionname+"1"].append(
                        [prevsubject] + ["".join(str(aa) for aa in str(row[4].value) if aa.isprintable())] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[6:] if cell.value is not None]
                    )
                    data[x][prev][sectionname+"2"].append(
                        [prevsubject] + ["".join(str(aa) for aa in str(row[5].value) if aa.isprintable())] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[6:] if cell.value is not None]
                    )
                else:
                    data[x][prev] = {}
                    data[x][prev][sectionname+"1"] = [
                        [prevsubject] + ["".join(str(aa) for aa in str(row[4].value) if aa.isprintable())] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[6:] if cell.value is not None]
                    ]
                    data[x][prev][sectionname+"2"] = [
                        [prevsubject] + ["".join(str(aa) for aa in str(row[5].value) if aa.isprintable())] + ["".join(str(aa) for aa in str(cell.value) if aa.isprintable()) for cell in row[6:] if cell.value is not None]
                    ]

with open('./json_data/sem4.json', 'w') as f:
    json.dump(data, f)